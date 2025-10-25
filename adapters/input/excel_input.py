import openpyxl
from typing import Iterable, Optional

from automation_library.core.interfaces import BaseInput, TaskItem


def _col_letter_to_index(col_letter: str) -> int:
    idx = 0
    for ch in col_letter.upper():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


class AtlasCopcoExcelInput(BaseInput):
    """Input concreto que lê códigos a partir de um arquivo Excel.

    Parâmetros principais:
    - file_path: caminho para o arquivo .xlsx
    - sheet_name: nome da planilha (opcional; se None usa a ativa)
    - code_column: letra da coluna que contém o código (default 'A')
    - start_row: primeira linha a ler (default 2, pulando cabeçalho)
    - id_key: nome da chave no TaskItem para o identificador (default 'code')
    - desc_column: letra da coluna com descrição (opcional)

    Esta implementação é lazy: `get_items` gera itens um a um usando
    `openpyxl.load_workbook(..., read_only=True)`.
    """

    uses_file: bool = True

    def __init__(self, file_path: str, sheet_name: Optional[str] = None, code_column: str = 'A', start_row: int = 2, id_key: str = 'code', desc_column: Optional[str] = None):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.code_column = code_column
        self.start_row = start_row
        self.id_key = id_key
        self.desc_column = desc_column

        self._wb = None
        self._sheet = None

    def open(self):
        self._wb = openpyxl.load_workbook(self.file_path, read_only=True)
        if self.sheet_name and self.sheet_name in self._wb.sheetnames:
            self._sheet = self._wb[self.sheet_name]
        else:
            self._sheet = self._wb.active

    def get_items(self) -> Iterable[TaskItem]:
        if not self._sheet:
            raise RuntimeError("Input not opened. Call open() before get_items().")

        code_idx = _col_letter_to_index(self.code_column)
        desc_idx = _col_letter_to_index(self.desc_column) if self.desc_column else None

        for row in self._sheet.iter_rows(min_row=self.start_row, values_only=True):
            try:
                raw_code = row[code_idx]
            except Exception:
                raw_code = None

            if raw_code is None or (isinstance(raw_code, str) and not raw_code.strip()):
                # pula linhas sem código
                continue

            code = str(raw_code).strip()
            # zfill curto (mantém compatibilidade com extractor que usava zfill(10))
            try:
                code = code.zfill(10)
            except Exception:
                pass

            item = {self.id_key: code}

            if desc_idx is not None:
                try:
                    item['description'] = row[desc_idx]
                except Exception:
                    item['description'] = None

            # Tentamos recuperar número da linha (openpyxl value-only não fornece row),
            # então não podemos obter row_num aqui — cliente pode prover se necessário.
            # Para compatibilidade, deixamos 'row_num' vazio quando não disponível.
            item['row_num'] = None

            yield item

    def close(self):
        if self._wb:
            try:
                self._wb.close()
            except Exception:
                pass

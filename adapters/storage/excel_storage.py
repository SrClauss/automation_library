
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import os
from typing import List, Dict, Any, Iterable
from datetime import datetime
import hashlib

from automation_library.core.interfaces import BaseStorage, DataItem, TaskItem

class ExcelStorage(BaseStorage):
    """
    Implementação de persistência para arquivos Excel (.xlsx).

    Gerencia a criação, atualização e continuação de processamentos em planilhas.
    """
    def __init__(self, output_file: str, sheet_name: str, headers: Dict[str, str], input_file_hash: str, task_id_key: str):
        self.output_file = output_file
        self.sheet_name = sheet_name
        self.headers_map = headers # Ex: {"code": "Código", "name": "Nome"}
        self.ordered_headers = list(headers.keys())
        self.input_file_hash = input_file_hash
        self.task_id_key = task_id_key # Chave que identifica unicamente a tarefa (ex: "row_num")
        self.workbook: Workbook = None
        self.data_sheet: Worksheet = None
        self.processed_ids = set()

    def open(self):
        """Carrega um workbook existente ou cria um novo."""
        if os.path.exists(self.output_file):
            self.workbook = openpyxl.load_workbook(self.output_file)
            if self.sheet_name in self.workbook.sheetnames:
                self.data_sheet = self.workbook[self.sheet_name]
            else:
                self._create_sheet()
        else:
            self.workbook = Workbook()
            self.data_sheet = self.workbook.active
            self.data_sheet.title = self.sheet_name
            self._write_headers()
        
        self._load_metadata()

    def _create_sheet(self):
        self.data_sheet = self.workbook.create_sheet(self.sheet_name)
        self._write_headers()

    def _write_headers(self):
        """Escreve os cabeçalhos na planilha com base no mapeamento."""
        header_titles = [self.headers_map.get(h, h) for h in self.ordered_headers]
        self.data_sheet.append(header_titles)

    def _load_metadata(self):
        """Carrega os metadados de um processamento anterior, se houver."""
        if "Metadata" in self.workbook.sheetnames:
            meta_sheet = self.workbook["Metadata"]
            stored_hash = meta_sheet.cell(row=1, column=2).value
            
            if stored_hash == self.input_file_hash:
                ids_str = meta_sheet.cell(row=2, column=2).value
                if ids_str:
                    self.processed_ids = {str(item_id) for item_id in str(ids_str).split(',')}

    def get_processed_items(self) -> Iterable[TaskItem]:
        """Retorna os IDs de tarefas já processados."""
        return self.processed_ids

    def save_items(self, items: List[DataItem]):
        """Salva uma lista de dicionários de dados na planilha."""
        if not items:
            return
        for item in items:
            task_id = item.get(self.task_id_key)
            if not task_id:
                continue # Ignora itens sem a chave de identificação

            # Constrói a linha na ordem correta dos cabeçalhos
            row_data = [item.get(h, "") for h in self.ordered_headers]

            # A chave da tarefa (ex: row_num) é usada como número da linha no Excel
            for col_idx, value in enumerate(row_data, start=1):
                self.data_sheet.cell(row=int(task_id), column=col_idx, value=value)

            self.processed_ids.add(str(task_id))

        # Após salvar o lote na memória, atualiza metadados e persiste imediatamente no disco
        try:
            self._update_metadata()
            if self.workbook:
                self.workbook.save(self.output_file)
        except Exception:
            # Não falha a aplicação em caso de erro ao salvar; apenas loga/ignora aqui
            pass

    def close(self):
        """Atualiza os metadados e salva o arquivo Excel."""
        self._update_metadata()
        if self.workbook:
            self.workbook.save(self.output_file)

    def _update_metadata(self):
        """Escreve/atualiza a planilha de metadados."""
        meta_sheet = self.workbook["Metadata"] if "Metadata" in self.workbook.sheetnames else self.workbook.create_sheet("Metadata")
        
        meta_sheet['A1'], meta_sheet['B1'] = "Input File Hash", self.input_file_hash
        meta_sheet['A2'], meta_sheet['B2'] = "Processed Item IDs", ",".join(sorted(list(self.processed_ids)))
        meta_sheet['A3'], meta_sheet['B3'] = "Timestamp", datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    @staticmethod
    def calculate_file_hash(filepath: str) -> str:
        """Calcula o hash SHA256 de um arquivo."""
        hash_sha256 = hashlib.sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

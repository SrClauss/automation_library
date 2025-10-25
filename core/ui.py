import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import openpyxl
import hashlib
import os
import json
import sys
import queue
import contextlib
import time
import psutil
from datetime import datetime
from typing import Type

try:
    from selenium.common.exceptions import WebDriverException, TimeoutException
except Exception:
    # Placeholders para permitir execução em ambientes sem selenium (testes com fakes)
    class WebDriverException(Exception):
        pass
    class TimeoutException(Exception):
        pass

from .interfaces import BaseAuthenticator, BaseExtractor, BaseStorage


def column_to_index(col_letter: str) -> int:
    index = 0
    for char in col_letter.upper():
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


class ScraperWorker(threading.Thread):
    def __init__(self, worker_id, headless_mode, login_event, authenticator: BaseAuthenticator, extractor: BaseExtractor, task_queue: queue.Queue, results_queue: queue.Queue, global_stop_event: threading.Event, login_log_queue: queue.Queue, scraper_log_queue: queue.Queue):
        super().__init__(daemon=True)
        self.worker_id = worker_id
        self.headless = headless_mode
        self.login_event = login_event
        self._stop_event = threading.Event()

        self.authenticator = authenticator
        self.extractor = extractor
        self.task_queue = task_queue
        self.results_queue = results_queue
        self.global_stop_event = global_stop_event
        self.login_log_queue = login_log_queue
        self.scraper_log_queue = scraper_log_queue

        if hasattr(self.extractor, 'log_queue'):
            self.extractor.log_queue = self.scraper_log_queue

    def stop(self):
        self._stop_event.set()

    def stopped(self):
        return self._stop_event.is_set()

    def _log_login(self, msg):
        self.login_log_queue.put(msg)

    def run(self):
        self._log_login(f"[Worker {self.worker_id}] Iniciando...")
        driver = None
        login_success = False

        try:
            self._log_login(f"[Worker {self.worker_id}] Tentando fazer login...")
            try:
                driver = self.authenticator.login()
            except TypeError:
                driver = self.authenticator.login()

            if driver:
                self._log_login(f"[Worker {self.worker_id}] ✅ Login bem-sucedido.")
                login_success = True
            else:
                self._log_login(f"[Worker {self.worker_id}] ❌ Falha no login.")

            self.login_event.set()

            if not login_success:
                return

            while not self.global_stop_event.is_set() and not self.stopped():
                try:
                    task = self.task_queue.get(timeout=1)
                    data = self.extractor.extract(driver, task)
                    self.results_queue.put(data)
                    self.task_queue.task_done()

                except queue.Empty:
                    continue
                except (WebDriverException, TimeoutException) as e:
                    self._log_login(f"🚨 [Worker {self.worker_id}] Erro no navegador: {type(e).__name__}. Reiniciando driver.")
                    self.task_queue.put(task)
                    if driver:
                        with contextlib.suppress(Exception): self.authenticator.logout(driver)
                    driver = None
                    while not driver and not self.global_stop_event.is_set() and not self.stopped():
                        self._log_login(f"[Worker {self.worker_id}] Retentando login...")
                        driver = self.authenticator.login()
                        if not driver:
                            time.sleep(30)
                    continue

        except Exception as e:
            self._log_login(f"🚨 [Worker {self.worker_id}] Erro crítico: {e}")
        finally:
            if driver:
                with contextlib.suppress(Exception): self.authenticator.logout(driver)
            self._log_login(f"[Worker {self.worker_id}] Finalizado.")


class FrameworkGUI(tk.Tk):
    def __init__(self, authenticator_class: Type[BaseAuthenticator], extractor_class: Type[BaseExtractor], storage_class: Type[BaseStorage], config: dict):
        super().__init__()
        self.title("Framework de Automação Genérico")
        self.geometry("1200x800")

        self.authenticator_class = authenticator_class
        self.extractor_class = extractor_class
        self.storage_class = storage_class
        self.config = config

        self.login_log_queue = queue.Queue()
        self.scraper_log_queue = queue.Queue()

        self.stop_event = threading.Event()
        self.tasks_queue = queue.Queue()
        self.results_queue = queue.Queue()
        self.unsaved_data = []
        self.worker_threads = []
        self.threads_lock = threading.Lock()

        self.total_items = 0
        self.saved_items_count = 0

        default_workers = self.config.get("engine_settings", {}).get("num_workers", 3)
        self.num_workers_var = tk.IntVar(value=default_workers)
        default_headless = self.config.get("engine_settings", {}).get("headless_mode", True)
        self.headless_var = tk.BooleanVar(value=default_headless)

        self.create_widgets()
        self.process_login_log_queue()
        self.process_scraper_log_queue()

    def log(self, message):
        self.login_log_queue.put(message)

    def create_widgets(self):
        file_frame = ttk.LabelFrame(self, text="Controle de Arquivos", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(file_frame, text="Selecionar Arquivo de Entrada",
                  command=self.select_input_file).pack(side=tk.LEFT)
        self.input_label = ttk.Label(file_frame, text="Nenhum arquivo selecionado")
        self.input_label.pack(side=tk.LEFT, padx=5)

        ttk.Button(file_frame, text="Selecionar Saída",
                  command=self.select_output_file).pack(side=tk.LEFT)
        self.output_label = ttk.Label(file_frame, text="Nenhum arquivo de saída definido")
        self.output_label.pack(side=tk.LEFT, padx=5)

        main_log_frame = ttk.Frame(self)
        main_log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        main_log_frame.columnconfigure(0, weight=1)
        main_log_frame.columnconfigure(1, weight=1)
        main_log_frame.rowconfigure(0, weight=1)

        left_log_frame = ttk.LabelFrame(main_log_frame, text="Logs de Login e Sistema", padding=5)
        left_log_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.login_log_area = scrolledtext.ScrolledText(left_log_frame, state='disabled', font=('Consolas', 10))
        self.login_log_area.pack(fill=tk.BOTH, expand=True)

        right_log_frame = ttk.LabelFrame(main_log_frame, text="Logs de Raspagem", padding=5)
        right_log_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        self.scraper_log_area = scrolledtext.ScrolledText(right_log_frame, state='disabled', font=('Consolas', 10))
        self.scraper_log_area.pack(fill=tk.BOTH, expand=True)

        progress_frame = ttk.Frame(self)
        progress_frame.pack(fill=tk.X, padx=10, pady=5)

        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_label = ttk.Label(progress_frame, text="0/0")
        self.progress_label.pack(side=tk.LEFT, padx=5)

        ctrl_frame = ttk.Frame(self)
        ctrl_frame.pack(fill=tk.X, padx=10, pady=5)

        self.start_btn = ttk.Button(ctrl_frame, text="INICIAR PROCESSAMENTO", command=self.start_process)
        self.start_btn.pack(side=tk.LEFT)

        self.stop_btn = ttk.Button(ctrl_frame, text="PARAR E SALVAR", command=self.stop_process, state='disabled')
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        ttk.Label(ctrl_frame, text="Workers:").pack(side=tk.LEFT, padx=(20, 2))
        self.workers_spinbox = ttk.Spinbox(ctrl_frame, from_=1, to=20, textvariable=self.num_workers_var, width=5)
        self.workers_spinbox.pack(side=tk.LEFT)

        self.headless_check = ttk.Checkbutton(ctrl_frame, text="Rodar em 2º plano (headless)", variable=self.headless_var, onvalue=True, offvalue=False)
        self.headless_check.pack(side=tk.LEFT, padx=(10, 0))

        self.status_var = tk.StringVar(value="Pronto para iniciar")
        ttk.Label(ctrl_frame, textvariable=self.status_var, font=('Arial', 10)).pack(side=tk.LEFT, padx=10)

        self.eta_var = tk.StringVar(value="ETA: --:--:--")
        ttk.Label(ctrl_frame, textvariable=self.eta_var, font=('Arial', 10, 'italic')).pack(side=tk.RIGHT, padx=10)

        self.speed_var = tk.StringVar(value="-- itens/min")
        ttk.Label(ctrl_frame, textvariable=self.speed_var, font=('Arial', 10, 'italic')).pack(side=tk.RIGHT, padx=5)

    def process_log_queue(self, log_queue, area):
        try:
            while True:
                message = log_queue.get_nowait()
                area.config(state='normal')
                area.insert(tk.END, str(message) + "\n")
                area.see(tk.END)
                area.config(state='disabled')
        except queue.Empty:
            pass
        finally:
            self.after(100, lambda: self.process_log_queue(log_queue, area))

    def process_login_log_queue(self):
        self.process_log_queue(self.login_log_queue, self.login_log_area)

    def process_scraper_log_queue(self):
        self.process_log_queue(self.scraper_log_queue, self.scraper_log_area)

    def select_input_file(self):
        file_path = filedialog.askopenfilename(title="Selecione o arquivo Excel de entrada", filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
        if file_path:
            try:
                self.input_file = file_path
                self.input_label.config(text=os.path.basename(file_path))
                self.input_hash = hashlib.sha256(open(file_path, "rb").read()).hexdigest()

                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                selected = self.ask_sheet_selection(sheets)
                if selected:
                    self.selected_sheet = selected
                    self.log(f"Planilha selecionada: {selected}")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível ler o arquivo:\n{str(e)}")

    def ask_sheet_selection(self, sheets):
        popup = tk.Toplevel(self)
        popup.title("Selecionar Planilha")
        popup.geometry("300x300")
        tk.Label(popup, text="Selecione a planilha para processar:").pack(pady=10)
        selected = tk.StringVar(value=sheets[0])
        for sheet in sheets:
            rb = tk.Radiobutton(popup, text=sheet, variable=selected, value=sheet, padx=20, pady=5)
            rb.pack(anchor='w')
        result = []
        def on_ok():
            result.append(selected.get())
            popup.destroy()
        tk.Button(popup, text="OK", command=on_ok).pack(pady=10)
        popup.grab_set()
        self.wait_window(popup)
        return result[0] if result else None

    def select_output_file(self):
        if not hasattr(self, 'input_file'):
            messagebox.showwarning("Aviso", "Selecione um arquivo de entrada primeiro.")
            return
        default_name = os.path.splitext(os.path.basename(self.input_file))[0] + "_PROCESSADO.xlsx"
        file_path = filedialog.asksaveasfilename(title="Salvar resultado como",defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")], initialfile=default_name)
        if file_path:
            self.output_file = file_path
            self.output_label.config(text=os.path.basename(file_path))
            self.log(f"Arquivo de saída definido para: {self.output_file}")

    def start_process(self):
        if not hasattr(self, 'input_file') or not hasattr(self, 'output_file') or not hasattr(self, 'selected_sheet'):
            messagebox.showwarning("Aviso", "Configure os arquivos de entrada, saída e a planilha.")
            return

        try:
            self.authenticator = self.authenticator_class(self.config['credentials'], self.headless_var.get(), log_queue=self.login_log_queue)
            self.extractor = self.extractor_class(log_queue=self.scraper_log_queue)

            self.storage = self.storage_class(
                output_file=self.output_file,
                sheet_name=self.selected_sheet,
                headers=self.config['excel_output']['headers'],
                input_file_hash=self.input_hash,
                task_id_key=self.config['excel_output']['task_id_key']
            )
            self.storage.open()
        except Exception as e:
            messagebox.showerror("Erro na Inicialização", f"Falha ao iniciar adaptadores: {e}")
            return

        processed_items = self.storage.get_processed_items()
        if processed_items:
            response = messagebox.askyesnocancel("Continuar Processamento?", f"Foram encontrados {len(processed_items)} itens já salvos.\n\nDeseja continuar de onde parou?", icon='question')
            if response is None:
                self.storage.close()
                return
            if not response:
                self.storage.processed_ids.clear()

        self.saved_items_count = len(self.storage.processed_ids)

        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.workers_spinbox.config(state='normal')
        self.headless_check.config(state='disabled')
        self.stop_event.clear()

        threading.Thread(target=self.run_scraping, daemon=True).start()

    def _worker_manager(self):
        """
        Gerencia o pool de workers com login em lotes e ajuste dinâmico.
        """
        worker_serial_id = 0
        login_batch_size = self.config.get("engine_settings", {}).get("login_batch_size", 3)
        self.log(f"MANAGER: Iniciando logins em lotes de {login_batch_size}.")

        while not self.stop_event.is_set():
            with self.threads_lock:
                # Remove threads que já terminaram da lista
                self.worker_threads = [t for t in self.worker_threads if t.is_alive()]
                
                target_workers = self.num_workers_var.get()
                current_workers = len(self.worker_threads)
                
                # Adiciona workers se necessário, em lotes
                if current_workers < target_workers:
                    # Calcula quantos workers iniciar neste lote
                    needed = target_workers - current_workers
                    batch_size = min(needed, login_batch_size)
                    self.log(f"MANAGER: Iniciando um lote de {batch_size} novo(s) worker(s).")
                    
                    login_events = []
                    for _ in range(batch_size):
                        worker_serial_id += 1
                        login_event = threading.Event()
                        login_events.append(login_event)
                        worker = ScraperWorker(worker_serial_id, self.headless_var.get(), login_event, self.authenticator, self.extractor, self.tasks_queue, self.results_queue, self.stop_event, self.login_log_queue, self.scraper_log_queue)
                        worker.start()
                        self.worker_threads.append(worker)

                    # Espera que todos os logins do lote terminem
                    self.log(f"MANAGER: Aguardando resultado do login do lote de {batch_size} worker(s)...")
                    for event in login_events:
                        event.wait()
                    self.log("MANAGER: Lote de logins concluído.")
                
                # Remove workers se necessário
                elif current_workers > target_workers:
                    to_remove = current_workers - target_workers
                    self.log(f"MANAGER: Sinalizando para remover {to_remove} worker(s).")
                    
                    # Pega os últimos workers da lista para parar
                    workers_to_stop = self.worker_threads[target_workers:]
                    for worker in workers_to_stop:
                        worker.stop()

            time.sleep(2)
        
        # Ao final do processo, sinaliza para todos os workers pararem
        with self.threads_lock:
            self.log("MANAGER: Sinal de parada global recebido. Encerrando todos os workers.")
            for worker in self.worker_threads:
                worker.stop()

    def run_scraping(self):
        try:
            self.log("\n=== INICIANDO PROCESSAMENTO ===")
            self.tasks_queue = queue.Queue()
            self.results_queue = queue.Queue()

            code_column_letter = self.config['excel_input']['code_column']
            self.log(f"Lendo códigos da coluna {code_column_letter}.")
            
            wb_input = openpyxl.load_workbook(self.input_file, read_only=True)
            sheet_input = wb_input[self.selected_sheet]
            
            all_valid_tasks = []
            code_col_idx = column_to_index(code_column_letter) - 1

            for row in sheet_input.iter_rows(min_row=2): 
                cell = row[code_col_idx]
                if cell.value and str(cell.value).strip():
                    all_valid_tasks.append({'code': str(cell.value).zfill(10), 'row_num': cell.row})

            wb_input.close()
            
            self.total_items = len(all_valid_tasks)
            tasks_to_run = [task for task in all_valid_tasks if str(task['row_num']) not in self.storage.processed_ids]
            
            self.log(f"{len(tasks_to_run)} tarefas novas para processar.")
            for task in tasks_to_run:
                self.tasks_queue.put(task)

            self.progress["maximum"] = len(tasks_to_run)
            self.progress_label.config(text=f"0/{len(tasks_to_run)}")
            
            manager_thread = threading.Thread(target=self._worker_manager, daemon=True)
            manager_thread.start()

            start_time = time.time()
            items_processed_session = 0
            while not self.stop_event.is_set():
                try:
                    data = self.results_queue.get(timeout=1)
                    if data:
                        self.unsaved_data.append(data)
                        
                        if len(self.unsaved_data) >= self.config.get("engine_settings", {}).get("save_interval", 15):
                            self.save_data()
                        
                        items_processed_session += 1
                        total_processed_overall = self.saved_items_count + items_processed_session

                        elapsed = time.time() - start_time
                        if elapsed > 2:
                            speed = items_processed_session / elapsed * 60
                            self.speed_var.set(f"{speed:.1f} itens/min")
                            remaining = len(tasks_to_run) - items_processed_session
                            if speed > 0 and remaining > 0:
                                eta_seconds = remaining / (speed / 60)
                                h, m, s = int(eta_seconds // 3600), int((eta_seconds % 3600) // 60), int(eta_seconds % 60)
                                self.eta_var.set(f"ETA: {h:02d}:{m:02d}:{s:02d}")
                        
                        self.progress_var.set(items_processed_session)
                        self.progress_label.config(text=f"{items_processed_session}/{len(tasks_to_run)}")
                        
                        with self.threads_lock:
                            active_workers = len([t for t in self.worker_threads if t.is_alive()])
                        target_workers = self.num_workers_var.get()
                        self.status_var.set(f"Processando... | Salvos: {total_processed_overall} | Workers: {active_workers}/{target_workers}")

                except queue.Empty:
                    if self.tasks_queue.empty():
                        with self.threads_lock:
                            active_workers = len([t for t in self.worker_threads if t.is_alive()])
                        target_workers = self.num_workers_var.get()
                        
                        if active_workers == 0 and target_workers == 0:
                            self.log("Fila vazia, nenhum worker ativo e target=0. Finalizando processamento.")
                            break
                        elif active_workers == 0:
                            self.log(f"Fila vazia mas target_workers={target_workers}. Aguardando workers ou novas tarefas...")
                            time.sleep(1)
                            continue
                    time.sleep(0.5)
            
            if self.unsaved_data:
                self.save_data()
            self.log("\nPROCESSAMENTO CONCLUÍDO." if not self.stop_event.is_set() else "\nProcessamento interrompido.")
            
        except Exception as e:
            self.log(f"\nERRO DURANTE PROCESSAMENTO: {e}")
            import traceback
            self.log(traceback.format_exc())
        finally:
            self.cleanup()
            final_status = "Concluído" if not self.stop_event.is_set() else "Interrompido"
            self.status_var.set(final_status)
            self.start_btn.config(state='normal')
            self.stop_btn.config(state='disabled')
            self.workers_spinbox.config(state='normal')
            self.headless_check.config(state='normal')

    def save_data(self):
        if not self.unsaved_data:
            return
        self.log(f"Salvando lote de {len(self.unsaved_data)} itens...")
        try:
            self.storage.save_items(self.unsaved_data)
            self.saved_items_count += len(self.unsaved_data)
            self.unsaved_data = []
            self.log(f"Lote salvo com sucesso.")
        except Exception as e:
            self.log(f"ERRO AO SALVAR: {e}")
            import traceback
            self.log(traceback.format_exc())
    
    def stop_process(self):
        self.stop_event.set()
        self.status_var.set("Finalizando...")
        self.log("\nSolicitação de parada recebida...")

    def cleanup(self):
        self.log("\nSinalizando para workers finalizarem...")
        if not self.stop_event.is_set():
            self.stop_event.set()
        with self.threads_lock:
            for thread in self.worker_threads:
                if thread.is_alive():
                    thread.join(timeout=5)
        
        if getattr(self, 'unsaved_data', None):
            try:
                self.save_data()
            except Exception:
                pass

        if hasattr(self, 'storage') and self.storage:
            try:
                self.storage.close()
            except Exception:
                pass

        try:
            for proc in psutil.process_iter(['pid', 'name']):
                if 'chrome' in proc.info['name'].lower():
                    self.log(f"Encerrando processo Chrome (PID: {proc.info['pid']})...")
                    proc.kill()
        except Exception as e:
            self.log(f"Erro ao limpar processos chrome: {e}")
        self.log("Limpeza concluída.")
    
    def on_closing(self):
        if messagebox.askokcancel("Sair", "Deseja realmente sair?"):
            self.stop_process()
            self.status_var.set("Finalizando...")
            threading.Thread(target=self._perform_cleanup_and_exit, daemon=True).start()

    def _perform_cleanup_and_exit(self):
        try:
            self.cleanup()
        except Exception:
            pass

        try:
            self.after(0, self.destroy)
        except Exception:
            try:
                self.destroy()
            except Exception:
                pass